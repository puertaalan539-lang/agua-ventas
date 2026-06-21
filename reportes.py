"""
reportes.py
Generación de reportes Excel diarios y semanales.
Librerías: pandas, openpyxl
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import date, timedelta

import database as db

OUTPUT_DIR = Path("reportes")
OUTPUT_DIR.mkdir(exist_ok=True)

# Colores corporativos
COLOR_HEADER  = "1A6B9A"   # azul agua
COLOR_SUBTOT  = "D6EAF8"   # azul claro
COLOR_TOTAL   = "154360"   # azul oscuro


# ──────────────────────────────────────────────
# HELPERS DE ESTILO
# ──────────────────────────────────────────────
def _estilo_header(ws, fila: int, col_ini: int, col_fin: int, texto: str, color=COLOR_HEADER):
    ws.merge_cells(
        start_row=fila, start_column=col_ini,
        end_row=fila,   end_column=col_fin
    )
    cell = ws.cell(row=fila, column=col_ini, value=texto)
    cell.fill      = PatternFill("solid", fgColor=color)
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _borde_fino():
    lado = Side(style="thin", color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _aplicar_borde_rango(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col,  max_col=max_col):
        for cell in row:
            cell.border = _borde_fino()


def _autoajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)


# ──────────────────────────────────────────────
# REPORTE DIARIO
# ──────────────────────────────────────────────
def generar_reporte_diario(fecha: date) -> Path:
    """
    Genera un Excel con:
    - Hoja 1: Ventas por local
    - Hoja 2: Ventas individuales
    - Hoja 3: Gastos y distribución
    """
    ruta = OUTPUT_DIR / f"reporte_diario_{fecha}.xlsx"

    ventas_loc  = db.get_ventas_por_fecha(fecha)
    ventas_ind  = db.get_ventas_individuales_fecha(fecha)
    gastos_list = db.get_gastos_fecha(fecha)

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        _hoja_ventas_locales(writer, ventas_loc, fecha)
        _hoja_ventas_individuales(writer, ventas_ind, fecha)
        _hoja_gastos(writer, gastos_list, ventas_loc, ventas_ind, fecha)

    _aplicar_formato_global(ruta)
    db.registrar_reporte("diario", str(ruta))
    return ruta


def _hoja_ventas_locales(writer, ventas, fecha):
    if not ventas:
        pd.DataFrame().to_excel(writer, sheet_name="Ventas Locales", index=False)
        return

    filas = []
    for v in ventas:
        filas.append({
            "Local":           v["local_nombre"],
            "Garrafón (un)":   v["garrafon"],
            "Medio Garrafón":  v["medio_garrafon"],
            "Galón":           v["galon"],
            "P. Garrafón $":   v["precio_garrafon"],
            "P. Medio $":      v["precio_medio"],
            "P. Galón $":      v["precio_galon"],
            "Total Bruto $":   round(v["total_bruto"], 2),
            "Fuente":          v["fuente"],
        })

    df = pd.DataFrame(filas)
    # Fila de totales
    totales = {
        "Local": "TOTAL",
        "Garrafón (un)":  df["Garrafón (un)"].sum(),
        "Medio Garrafón": df["Medio Garrafón"].sum(),
        "Galón":          df["Galón"].sum(),
        "Total Bruto $":  df["Total Bruto $"].sum(),
    }
    df = pd.concat([df, pd.DataFrame([totales])], ignore_index=True)
    df.to_excel(writer, sheet_name="Ventas Locales", index=False, startrow=2)


def _hoja_ventas_individuales(writer, ventas, fecha):
    if not ventas:
        pd.DataFrame(columns=["Fecha","Producto","Cantidad","Precio Unit $","Total $","Notas"])\
          .to_excel(writer, sheet_name="Ventas Individuales", index=False)
        return

    filas = [
        {
            "Fecha":          str(v["fecha"]),
            "Producto":       v["producto"].replace("_", " ").title(),
            "Cantidad":       v["cantidad"],
            "Precio Unit $":  v["precio_unit"],
            "Total $":        round(v["total"], 2),
            "Notas":          v["notas"],
        }
        for v in ventas
    ]
    df = pd.DataFrame(filas)
    totales = {"Fecha": "TOTAL", "Cantidad": df["Cantidad"].sum(),
               "Total $": df["Total $"].sum()}
    df = pd.concat([df, pd.DataFrame([totales])], ignore_index=True)
    df.to_excel(writer, sheet_name="Ventas Individuales", index=False, startrow=2)


def _hoja_gastos(writer, gastos, ventas_loc, ventas_ind, fecha):
    filas_g = []
    tot_garr = tot_med = tot_gal = 0.0

    for g in gastos:
        m = g["monto_total"]
        g_garr = round(m * g["porc_garrafon"], 2)
        g_med  = round(m * g["porc_medio"],    2)
        g_gal  = round(m * g["porc_galon"],    2)
        tot_garr += g_garr; tot_med += g_med; tot_gal += g_gal
        filas_g.append({
            "Descripción": g["descripcion"],
            "Monto Total $": m,
            "% Garrafón":  f"{g['porc_garrafon']*100:.0f}%",
            "Gasto Garrafón $": g_garr,
            "% Medio":     f"{g['porc_medio']*100:.0f}%",
            "Gasto Medio $":    g_med,
            "% Galón":     f"{g['porc_galon']*100:.0f}%",
            "Gasto Galón $":    g_gal,
        })

    # Ventas brutas
    bruto_loc = sum(v["total_bruto"] for v in ventas_loc)
    bruto_ind = sum(v["total"]       for v in ventas_ind)
    bruto_tot = bruto_loc + bruto_ind
    gasto_tot = tot_garr + tot_med + tot_gal
    neto      = round(bruto_tot - gasto_tot, 2)

    resumen = {
        "Descripción":         "RESUMEN DÍA",
        "Ventas Locales $":    round(bruto_loc, 2),
        "Ventas Individuales $": round(bruto_ind, 2),
        "Total Bruto $":       round(bruto_tot, 2),
        "Total Gastos $":      round(gasto_tot, 2),
        "TOTAL NETO DÍA $":   neto,
    }

    df_g = pd.DataFrame(filas_g) if filas_g else pd.DataFrame()
    df_r = pd.DataFrame([resumen])

    df_g.to_excel(writer, sheet_name="Gastos", index=False, startrow=2)
    df_r.to_excel(writer, sheet_name="Resumen", index=False, startrow=2)


def _aplicar_formato_global(ruta: Path):
    wb = load_workbook(ruta)
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        _estilo_header(ws, 1, 1, ws.max_column or 1,
                       f"AGUA PURIFICADA — {nombre_hoja.upper()}")
        # Encabezados de columnas (fila 3)
        for cell in ws[3]:
            if cell.value:
                cell.fill      = PatternFill("solid", fgColor=COLOR_SUBTOT)
                cell.font      = Font(bold=True, color="154360")
                cell.alignment = Alignment(horizontal="center")
        # Última fila en negrita (totales)
        if ws.max_row > 3:
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EBF5FB")
        _aplicar_borde_rango(ws, 3, ws.max_row, 1, ws.max_column or 1)
        _autoajustar_columnas(ws)
    wb.save(ruta)


# ──────────────────────────────────────────────
# REPORTE SEMANAL
# ──────────────────────────────────────────────
def generar_reporte_semanal(fecha_fin: date) -> Path:
    """
    Agrupa los datos de los últimos 7 días y genera un Excel ejecutivo.
    """
    fecha_ini = fecha_fin - timedelta(days=6)
    ruta = OUTPUT_DIR / f"reporte_semanal_{fecha_ini}_{fecha_fin}.xlsx"

    ventas = db.get_ventas_rango(fecha_ini, fecha_fin)

    # Agrupar por fecha
    resumen: dict[str, dict] = {}
    for v in ventas:
        f = str(v["fecha"])
        if f not in resumen:
            resumen[f] = {"Fecha": f, "Garrafón": 0, "Medio Garrafón": 0,
                          "Galón": 0, "Total Bruto $": 0.0}
        resumen[f]["Garrafón"]       += v["garrafon"]
        resumen[f]["Medio Garrafón"] += v["medio_garrafon"]
        resumen[f]["Galón"]          += v["galon"]
        resumen[f]["Total Bruto $"]  += v["total_bruto"]

    filas = list(resumen.values())
    if filas:
        totales = {
            "Fecha":          "TOTAL SEMANA",
            "Garrafón":       sum(f["Garrafón"]       for f in filas),
            "Medio Garrafón": sum(f["Medio Garrafón"] for f in filas),
            "Galón":          sum(f["Galón"]          for f in filas),
            "Total Bruto $":  round(sum(f["Total Bruto $"] for f in filas), 2),
        }
        filas.append(totales)

    df = pd.DataFrame(filas)

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resumen Semanal", index=False, startrow=2)

        # Hoja por local
        locales = db.get_locales()
        for local in locales:
            filas_l = [
                {
                    "Fecha":          str(v["fecha"]),
                    "Garrafón":       v["garrafon"],
                    "Medio Garrafón": v["medio_garrafon"],
                    "Galón":          v["galon"],
                    "Total $":        round(v["total_bruto"], 2),
                }
                for v in ventas if v["local_id"] == local["id"]
            ]
            if filas_l:
                pd.DataFrame(filas_l).to_excel(
                    writer, sheet_name=local["nombre"][:30], index=False, startrow=2
                )

    _aplicar_formato_global(ruta)
    db.registrar_reporte("semanal", str(ruta))
    return ruta